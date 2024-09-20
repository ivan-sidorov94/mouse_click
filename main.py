import pyautogui
import subprocess
import os
import time
import configparser
import pandas as pd
import babel.numbers
import tzdata
import tkinter as tk
import shutil

from tkinter import filedialog
from openpyxl.styles import Border, Side
from openpyxl.worksheet.table import Table
from datetime import datetime, timedelta
from tkcalendar import DateEntry
from UliPlot.XLSX import auto_adjust_xlsx_column_width

path_to_config = "C:\\Program Files\\EleSy\\SCADA Infinity\\InfinityAlarms\\settings.ini"

# path_to_InfinityAlarms = "C:\\Program Files (x86)\\EleSy\\SCADA Infinity\\InfinityAlarms\\InfinityAlarms.exe"

# path_to_raw_file = "C:\\Program Files (x86)\\EleSy\\SCADA Infinity\\InfinityAlarms"

default_path_to_save_ASUTP = "D:\\Alarms_ASUTP"
default_path_to_save_ASUE = "D:\\Alarms_ASUE"
default_path_to_save_TVS = "D:\\Alarms_TVS"
default_path_to_save_ASPS = "D:\\Alarms_ASPS"

config_file = path_to_config
config = configparser.RawConfigParser()

if not os.path.exists(config_file):
    config.add_section('Time')
    config.set('Time', 'time1', '45')
    config.set('Time', 'time2', '0.1')
    config.set('Time', 'time3', '20')
    config.add_section('Start')
    config.set('Start', 'autostart', 'False')
    config.add_section('Save Directory')
    config.set('Save Directory', 'save_folder_path_ASUTP', f'{default_path_to_save_ASUTP}')
    config.set('Save Directory', 'save_folder_path_ASUE', f'{default_path_to_save_ASUE}')
    config.set('Save Directory', 'save_folder_path_TVS', f'{default_path_to_save_TVS}')
    config.set('Save Directory', 'save_folder_path_ASPS', f'{default_path_to_save_ASPS}')
    config.add_section('IPs')
    config.set('IPs', 'Server_ASUTP_IP', '200.1.1.1')
    config.set('IPs', 'Server_ASUE_IP', '200.1.1.16')
    config.set('IPs', 'Server_TVS_IP', '200.1.1.11')
    config.set('IPs', 'Server_ASPS_IP', '200.1.1.3')
    config.add_section('Path')
    config.set('Path', 'Path_to_InfinityAlarms', 'C:\\Program Files\\EleSy\\SCADA Infinity\\InfinityAlarms\\InfinityAlarms.exe')
    config.set('Path', 'Path_to_Raw_File', 'C:\\Program Files\\EleSy\\SCADA Infinity\\InfinityAlarms')
    config.add_section('System')
    config.set('System', 'ASUTP', 'True')
    config.set('System', 'ASUE', 'False')
    config.set('System', 'TVS', 'False')
    config.set('System', 'ASPS', 'False')

    with open(config_file, 'w') as configfile:
        config.write(configfile)

try:
    config.read(config_file)

    time1 = config.getfloat('Time', 'time1')
    time2 = config.getfloat('Time', 'time2')
    time3 = config.getfloat('Time', 'time3')
    autostart = config.getboolean('Start', 'autostart')
    save_folder_ASUTP = config.get('Save Directory', 'save_folder_path_ASUTP')
    save_folder_ASUE = config.get('Save Directory', 'save_folder_path_ASUE')
    save_folder_TVS = config.get('Save Directory', 'save_folder_path_TVS')
    save_folder_ASPS = config.get('Save Directory', 'save_folder_path_ASPS')
    Server_ASUTP_IP = config.get('IPs', 'Server_ASUTP_IP')
    Server_ASUE_IP = config.get('IPs', 'Server_ASUE_IP')
    Server_TVS_IP = config.get('IPs', 'Server_TVS_IP')
    Server_ASPS_IP = config.get('IPs', 'Server_ASPS_IP')
    path_to_InfinityAlarms = config.get('Path', 'Path_to_InfinityAlarms')
    path_to_raw_file = config.get('Path', 'Path_to_Raw_File')
    get_ASUTP = config.getboolean('System', 'ASUTP')
    get_ASUE = config.getboolean('System', 'ASUE')
    get_TVS = config.getboolean('System', 'TVS')
    get_ASPS = config.getboolean('System', 'ASPS')
    


except :
    time1 = 45
    time2 = 0.5
    time3 = 45
    config.clear()
    config.add_section('Time')
    config.set('Time', 'time1', '45')
    config.set('Time', 'time2', '0.1')
    config.set('Time', 'time3', '20')
    config.add_section('Start')
    config.set('Start', 'autostart', 'False')
    config.add_section('Save Directory')
    config.set('Save Directory', 'save_folder_path_ASUTP', f'{default_path_to_save_ASUTP}')
    config.set('Save Directory', 'save_folder_path_ASUE', f'{default_path_to_save_ASUE}')
    config.set('Save Directory', 'save_folder_path_TVS', f'{default_path_to_save_TVS}')
    config.set('Save Directory', 'save_folder_path_ASPS', f'{default_path_to_save_ASPS}')
    config.add_section('IPs')
    config.set('IPs', 'Server_ASUTP_IP', '200.1.1.1')
    config.set('IPs', 'Server_ASUE_IP', '200.1.1.16')
    config.set('IPs', 'Server_TVS_IP', '200.1.1.11')
    config.set('IPs', 'Server_ASPS_IP', '200.1.1.3')
    config.add_section('Path')
    config.set('Path', 'Path_to_InfinityAlarms', 'C:\\Program Files\\EleSy\\SCADA Infinity\\InfinityAlarms\\InfinityAlarms.exe')
    config.set('Path', 'Path_to_Raw_File', 'C:\\Program Files\\EleSy\\SCADA Infinity\\InfinityAlarms\\')
    config.add_section('System')
    config.set('System', 'ASUTP', 'True')
    config.set('System', 'ASUE', 'False')
    config.set('System', 'TVS', 'False')
    config.set('System', 'ASPS', 'False')

    with open(config_file, 'w') as configfile:
            config.write(configfile)

    time1 = config.getfloat('Time', 'time1')
    time2 = config.getfloat('Time', 'time2')
    time3 = config.getfloat('Time', 'time3')
    autostart = config.getboolean('Start', 'autostart')
    save_folder_ASUTP = config.get('Save Directory', 'save_folder_path_ASUTP')
    save_folder_ASUE = config.get('Save Directory', 'save_folder_path_ASUE')
    save_folder_TVS = config.get('Save Directory', 'save_folder_path_TVS')
    save_folder_ASPS = config.get('Save Directory', 'save_folder_path_ASPS')
    Server_ASUTP_IP = config.get('IPs', 'Server_ASUTP_IP')
    Server_ASUE_IP = config.get('IPs', 'Server_ASUE_IP')
    Server_TVS_IP = config.get('IPs', 'Server_TVS_IP')
    Server_ASPS_IP = config.get('IPs', 'Server_ASPS_IP')
    path_to_InfinityAlarms = config.get('Path', 'Path_to_InfinityAlarms')
    path_to_raw_file = config.get('Path', 'Path_to_Raw_File')
    get_ASUTP = config.getboolean('System', 'ASUTP')
    get_ASUE = config.getboolean('System', 'ASUE')
    get_TVS = config.getboolean('System', 'TVS')
    get_ASPS = config.getboolean('System', 'ASPS')

def open_file():
    file_path = filedialog.askopenfilename(title="Выбор файла", initialdir=f"{path_to_InfinityAlarms}", filetypes=(("exe files","*.exe"),("all files","*.*")), initialfile="InfinityAlarms.exe")
    if file_path != "":
        file_path = file_path.replace('/', '\\')
        file_dir = os.path.dirname(file_path)
        config.set("Path", "Path_to_InfinityAlarms", f"{file_path}")
        config.set("Path", "Path_to_Raw_File", f"{file_dir}")
        with open(config_file, 'w') as f:
              config.write(f)


def alarms(date_format, int1, int2, int3, date1, date2, system_name, ip):
    date1 = datetime.strftime(date1, date_format)
    date2 = datetime.strftime(date2, date_format)

    config.set("Time", "time1", int1)
    config.set("Time", "time2", int2)
    config.set("Time", "time3", int3)
    match system_name:
        case 'ASUTP':
            config.set("IPs", 'Server_ASUTP_IP', ip)
        case 'ASUE':
            config.set("IPs", 'Server_ASUE_IP', ip)
        case 'TVS':
            config.set("IPs", 'Server_TVS_IP', ip)
        case 'ASPS':
            config.set("IPs", 'Server_ASPS_IP', ip)


    with open(config_file, 'w') as configfile:
        config.write(configfile)

    cmd = (f'{path_to_InfinityAlarms} HISTORY DBEG="{date1}" DEND"={date2}" '
           'Bookmark="Новая закладка"')
    p = subprocess.Popen(cmd)
    time.sleep(2)
    pyautogui.press('enter')
    time.sleep(int1)

    pyautogui.click(x=2065, y=32)
    pyautogui.sleep(int2)

    pyautogui.click(x=2066, y=59)
    pyautogui.sleep(int2)

    pyautogui.click(x=2575, y=872)
    pyautogui.sleep(int2)

    pyautogui.click(x=2609, y=456)
    pyautogui.sleep(int2)

    pyautogui.click(x=3025, y=439)
    pyautogui.sleep(int2)

    pyautogui.press('backspace', presses=10, interval=0.25)
    pyautogui.write(ip)
    # match system_name:
    #     case 'ASUTP':
    #         pyautogui.write(Server_ASUTP_IP)
    #     case 'ASUE':
    #         pyautogui.write(Server_ASUE_IP)
    #     case 'TVS':
    #         pyautogui.write(Server_TVS_IP)
    #     case 'ASPS':
    #         pyautogui.write(Server_ASPS_IP)
    pyautogui.sleep(int2)

    pyautogui.click(x=3146, y=868)
    pyautogui.sleep(int2)

    pyautogui.click(x=3146, y=617)

    time.sleep(2)
    p.kill()
    p = subprocess.Popen(cmd)
    time.sleep(2)

    pyautogui.press('enter')

    time.sleep(int1)
    pyautogui.click(x=1939, y=31)
    pyautogui.sleep(int2)
    pyautogui.click(x=1960, y=74)
    pyautogui.sleep(int2)
    pyautogui.click(x=2667, y=582)
    pyautogui.sleep(int2)
    pyautogui.click(x=2980, y=745)
    pyautogui.sleep(int2)
    pyautogui.click(x=2958, y=689)
    pyautogui.sleep(int3)
    # pyautogui.click(x=3050, y=623)
    p.kill()


def open_directory(system_name):
    save_folder_path = config.get('Save Directory', f'save_folder_path_{system_name}')
    try:
        os.startfile(save_folder_path)
    except FileNotFoundError:
        os.makedirs(save_folder_path, exist_ok=True)
        os.startfile(save_folder_path)


def obrabotka(date1, date2, system_name):

    date_format = '%Y_%m_%d'
    date1 = datetime.strftime(date1, date_format)
    date2 = datetime.strftime(date2, '%Y.%m.%d')

    file_name = rf'{path_to_raw_file}\Alarms_{date1}.xls'

    match system_name:
        case 'ASUTP':
            path_to_rem_file = f'{save_folder_ASUTP}\\Ремонт\\{date2}_Ремонты.xls'
        case 'ASUE':
            path_to_rem_file = f'{save_folder_ASUE}\\Ремонт\\{date2}_Ремонты.xls'
        case 'TVS':
            path_to_rem_file = f'{save_folder_TVS}\\Ремонт\\{date2}_Ремонты.xls'
        case 'ASPS':
            path_to_rem_file = f'{save_folder_ASPS}\\Ремонт\\{date2}_Ремонты.xls'

    sheet_name = 'Лист1'

    no_rem = False

    # Читаем только необходимые столбцы
    df = pd.read_excel(file_name, sheet_name=sheet_name, usecols=[0, 1, 2, 4, 5],
                       skiprows=2, names=['Время','Сообщение', 'Класс сообщения',
                                          'Состояние', 'Мнемосхема'], decimal=',', dtype=str)

    # Заменяем NaN на пустую строку
    df = df.fillna('')


    try:
        # Создаем новый фрейм для ремонта
        df_rem = df[df['Сообщение'].str.contains('ремонт"') | df['Сообщение'].str.contains('ремонта"')]
        df_rem = df_rem[~df_rem['Состояние'].str.contains('не')]

        # Удаляем ненужные столбцы из фреймов
        df_rem = df_rem.drop(columns=['Класс сообщения','Состояние','Мнемосхема'])

        # Создаем временный фрейм для обработки сообщений о ремонте, разбивая столбец Сообщение на 2
        new_df = df_rem['Сообщение'].str.split(' Под', expand=True, n=-1)

        # Удаляем столбец перед конкатенацией
        df_rem = df_rem.drop(columns='Сообщение')

        # Соединяем оба фрейма
        df_rem = pd.concat([df_rem, new_df], axis=1)

        # Удаляем из памяти временный фрейм
        del new_df

        # Переименовываем столбцы
        df_rem.columns = ['Время','Датчик|ИМ','Команда']

        # Изменяем значения в столбце Команда в зависимости от значения
        df_rem.loc[df_rem['Команда'].str.contains('Снять'), 'Команда'] =  'Снят с ремонта'
        df_rem.loc[df_rem['Команда'].str.contains('Вывести'), 'Команда'] =  'Выведен в ремонт'
    except:
        no_rem = True
        pass

    df = df.drop(columns=['Время'])
# Подсчитываем дубликаты напрямую
    df = df.groupby(list(df.columns)).size().reset_index(name='Dublicates')

# Сортируем по количеству дубликатов
    df = df.sort_values(by='Dublicates', ascending=False)

# Считаем общее количество
    total_dub = df['Dublicates'].sum()

# Добавляем строку с общим количеством
    df.loc[len(df.index)] = ['', '', '','Всего сообщений:',total_dub]
# Создаем объект Border с границами
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))

    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='NoDublicates', index=False)

    # Получаем объект workbook
        workbook = writer.book

# Получаем объект worksheet для другого листа
        worksheet = writer.sheets['Лист1']

# Получаем номер последней строки с данными
        num_rows = worksheet.max_row

# Создаем объект Table для диапазона данных на другом листе, начиная с 3 строки
        last_col_letter = 'F'  # Замените на букву последнего столбца
        tab = Table(displayName="Table1",ref="A3:" + last_col_letter + str(num_rows))

# Добавляем таблицу в другой лист
        worksheet.add_table(tab)

# Получаем объект worksheet
        worksheet = writer.sheets['NoDublicates']

        for row in worksheet.iter_rows():
            for cell in row:
# Применяем границы к каждой ячейке
                cell.border = thin_border

# Создаем объект Table для всего диапазона данных и добавляем фильтр
        tab = Table(displayName="Table2",ref="A1:E" + str(len(df.index) + 1))

# Добавляем таблицу в лист
        worksheet.add_table(tab)

# Автоматическое расширение столбцов
        auto_adjust_xlsx_column_width(df, writer, sheet_name="NoDublicates", index=False)

        if not no_rem:
            with pd.ExcelWriter(path_to_rem_file, engine='openpyxl', mode='w') as writer:
                df_rem.to_excel(writer, sheet_name='Ремонт', index=False)

                worksheet = writer.sheets['Ремонт']

                for row in worksheet.iter_rows():
                    for cell in row:
                    # Применяем границы к каждой ячейке
                        cell.border = thin_border

                    # Создаем объект Table для всего диапазона данных и добавляем фильтр
                tab = Table(displayName="Table3",ref="A1:C" + str(len(df.index) + 1))

                    # Добавляем таблицу в лист
                worksheet.add_table(tab)

                auto_adjust_xlsx_column_width(df_rem, writer, sheet_name="Ремонт", index=False)

    match system_name:
        case 'ASUTP':
            shutil.move(file_name, f'{save_folder_ASUTP}\\Отчет по алармам за период_{date2}.xls')
        case 'ASUE':
            shutil.move(file_name, f'{save_folder_ASUE}\\Отчет по алармам за период_{date2}.xls')
        case 'TVS':
            shutil.move(file_name, f'{save_folder_TVS}\\Отчет по алармам за период_{date2}.xls')
        case 'ASPS':
            shutil.move(file_name, f'{save_folder_ASPS}\\Отчет по алармам за период_{date2}.xls')

    del df
    try:
        del df_rem
    except:
        pass


    if var and system_name == 'ASPS':
        root.destroy()


def autostart_cmd():

    if asutp.get() == 1:
        config.set('System', 'ASUTP', 'True')
    else:
        config.set('System', 'ASUTP', 'False')
    with open(config_file, 'w') as f:
        config.write(f)


    if asue.get() == 1:
        config.set('System', 'ASUE', 'True')
    else:
        config.set('System', 'ASUE', 'False')
    with open(config_file, 'w') as f:
        config.write(f)


    if tvs.get() == 1:
        config.set('System', 'TVS', 'True')
    else:
        config.set('System', 'TVS', 'False')
    with open(config_file, 'w') as f:
        config.write(f)


    if asps.get() == 1:
        config.set('System', 'ASPS', 'True')
    else:
        config.set('System', 'ASPS', 'False')
    with open(config_file, 'w') as f:
        config.write(f)

    if var.get() == 1:
        config.set('Start', 'autostart', 'True')
    else:
        config.set('Start', 'autostart', 'False')
    with open(config_file, 'w') as f:
        config.write(f)


def set_save_folder(system_name):
    save_folder_path = filedialog.askdirectory()
    config.set('Save Directory', f'save_folder_path_{system_name}', f'{save_folder_path}')
    with open(config_file, 'w') as f:
        config.write(f)

def start(date_format, int1, int2, int3, date_entry1, date_entry2):
    if get_ASUTP:
        alarms(date_format, int1, int2, int3, date_entry1, date_entry2, system_name='ASUTP', ip=Server_ASUTP_IP)
        time.sleep(5)
        obrabotka(date_entry1, date_entry2, system_name='ASUTP')

    if get_ASUE:
        alarms(date_format, int1, int2, int3, date_entry1, date_entry2, system_name='ASUE', ip=Server_ASUE_IP)
        time.sleep(5)
        obrabotka(date_entry1, date_entry2, system_name='ASUE')

    if get_TVS:
        alarms(date_format, int1, int2, int3, date_entry1, date_entry2, system_name='TVS', ip=Server_TVS_IP)
        time.sleep(5)
        obrabotka(date_entry1, date_entry2, system_name='TVS')

    if get_ASPS:
        alarms(date_format, int1, int2, int3, date_entry1, date_entry2, system_name='ASPS', ip=Server_ASPS_IP)
        time.sleep(5)
        obrabotka(date_entry1, date_entry2, system_name='ASPS')


def main():

    # Вычисление вчерашней и позавчерашней даты
    date_format = '%d.%m.%Y'
    # before_yesterday = datetime.now() - timedelta(days=2)
    yesterday = datetime.now() - timedelta(days=1)
    today = datetime.now()
    # before_yesterday = before_yesterday.strftime(date_format)
    yesterday = yesterday.strftime(date_format)
    today = today.strftime(date_format)

    global var
    global asutp
    global asue
    global tvs
    global asps
    global root

    root = tk.Tk()
    var = tk.IntVar(value=config.getboolean('Start', 'autostart'))
    asutp = tk.IntVar(value=config.getboolean('System', 'ASUTP'))
    asue = tk.IntVar(value=config.getboolean('System', 'ASUE'))
    tvs = tk.IntVar(value=config.getboolean('System', 'TVS'))
    asps = tk.IntVar(value=config.getboolean('System', 'ASPS'))
    root.title("Выгрузка алармов за предыдущий день")

    frame1 = tk.Frame(root)
    frame2 = tk.Frame(root)
    frame3 = tk.Frame(root)

    # Создание меток и полей выбора даты
    label1 = tk.Label(frame1, text="Выберите начальную дату:")
    date_entry1 = DateEntry(frame1, date_pattern='dd.mm.yyyy')
    date_entry1.set_date(yesterday)

    label2 = tk.Label(frame1, text="Выберите конечную дату:")
    date_entry2 = DateEntry(frame1, date_pattern='dd.mm.yyyy')
    date_entry2.set_date(today)

    # Создание меток и полей ввода
    label3 = tk.Label(frame1, text="Введите задержку открытия программы:")
    int1 = tk.Entry(frame1)
    int1.insert(0, str(time1))

    label4 = tk.Label(frame1, text="Введите задержку перемещения мыши:")
    int2 = tk.Entry(frame1)
    int2.insert(0, str(time2))

    label5 = tk.Label(frame1, text="Введите задержку перед сохранением:")
    int3 = tk.Entry(frame1)
    int3.insert(0, str(time3))

    label6 = tk.Label(frame2, text="Введите IP Сервера АСУ ТП:")
    int4 = tk.Entry(frame2)
    int4.insert(0, Server_ASUTP_IP)
    check_button_ASUTP = tk.Checkbutton(frame2, text='Автостарт АСУ ТП', variable=asutp,
                                  command=autostart_cmd)

    label7 = tk.Label(frame2, text="Введите IP Сервера АСУЭ:")
    int5 = tk.Entry(frame2)
    int5.insert(0, Server_ASUE_IP)
    check_button_ASUE = tk.Checkbutton(frame2, text='Автостарт АСУЭ', variable=asue,
                                  command=autostart_cmd)

    label8 = tk.Label(frame2, text="Введите IP Сервера ТВС:")
    int6 = tk.Entry(frame2)
    int6.insert(0, Server_TVS_IP)
    check_button_TVS = tk.Checkbutton(frame2, text='Автостарт ТВС', variable=tvs,
                                  command=autostart_cmd)

    label9 = tk.Label(frame2, text="Введите IP Сервера АСПС:")
    int7 = tk.Entry(frame2)
    int7.insert(0, Server_ASPS_IP)
    check_button_ASPS = tk.Checkbutton(frame2, text='Автостарт АСПС', variable=asps,
                                  command=autostart_cmd)


    dir_button_ASUTP = tk.Button(frame2, text="Выбрать папку для сохранения АСУ ТП",
                           command=lambda: set_save_folder(system_name='ASUTP'))
    dir_button_ASUE = tk.Button(frame2, text="Выбрать папку для сохранения АСУЭ",
                           command=lambda: set_save_folder(system_name='ASUE'))
    dir_button_TVS = tk.Button(frame2, text="Выбрать папку для сохранения ТВС",
                           command=lambda: set_save_folder(system_name='TVS'))
    dir_button_ASPS = tk.Button(frame2, text="Выбрать папку для сохранения АСПС",
                           command=lambda: set_save_folder(system_name='ASPS'))


    # # Создание кнопок
    button1 = tk.Button(frame2, text="Запустить Infinity Alarms АСУ ТП", command=lambda:
                        alarms(date_format, float(int1.get()), float(int2.get()),
                               float(int3.get()), date_entry1.get_date(),
                               date_entry2.get_date(), system_name='ASUTP', ip=str(int4.get())))
    button2 = tk.Button(frame2, text="Запустить Infinity Alarms АСУЭ", command=lambda:
                        alarms(date_format, float(int1.get()), float(int2.get()),
                               float(int3.get()), date_entry1.get_date(),
                               date_entry2.get_date(), system_name='ASUE', ip=str(int5.get())))
    button3 = tk.Button(frame2, text="Запустить Infinity Alarms ТВС", command=lambda:
                        alarms(date_format, float(int1.get()), float(int2.get()),
                               float(int3.get()), date_entry1.get_date(),
                               date_entry2.get_date(), system_name='TVS', ip=str(int6.get())))
    button4 = tk.Button(frame2, text="Запустить Infinity Alarms АСПС", command=lambda:
                        alarms(date_format, float(int1.get()), float(int2.get()),
                               float(int3.get()), date_entry1.get_date(),
                               date_entry2.get_date(), system_name='ASPS', ip=str(int7.get())))

    button5 = tk.Button(frame2, text="Открыть каталог АСУ ТП", command=lambda: open_directory(system_name='ASUTP'))
    button6 = tk.Button(frame2, text="Открыть каталог АСУЭ", command=lambda: open_directory(system_name='ASUE'))
    button7 = tk.Button(frame2, text="Открыть каталог ТВС", command=lambda: open_directory(system_name='TVS'))
    button8 = tk.Button(frame2, text="Открыть каталог АСПС", command=lambda: open_directory(system_name='ASPS'))

    button9 = tk.Button(frame2, text="Запустить обработку АСУ ТП", command=lambda:
                        obrabotka(date_entry1.get_date(),
                                  date_entry2.get_date(),
                                  system_name='ASUTP'))
    button10 = tk.Button(frame2, text="Запустить обработку АСУЭ", command=lambda:
                        obrabotka(date_entry1.get_date(),
                                  date_entry2.get_date(),
                                  system_name='ASUE'))
    button11 = tk.Button(frame2, text="Запустить обработку ТВС", command=lambda:
                        obrabotka(date_entry1.get_date(),
                                  date_entry2.get_date(),
                                  system_name='TVS'))
    button12 = tk.Button(frame2, text="Запустить обработку АСПС", command=lambda:
                        obrabotka(date_entry1.get_date(),
                                  date_entry2.get_date(),
                                  system_name='ASPS'))


    fileopen_button = tk.Button(frame3, text="Указать путь до InfinityAlarms", command=lambda: open_file())

    start_button = tk.Button(frame3, text="Запустить всё", command=lambda: start(date_format, float(int1.get()), float(int2.get()),
                               float(int3.get()), date_entry1.get_date(), date_entry2.get_date()))

    check_button = tk.Checkbutton(frame3, text='Автостарт', variable=var,
                                  command=autostart_cmd)

    # Размещение меток и полей ввода на экране
    label1.grid(column=1, row=1)
    label2.grid(column=2, row=1)
    date_entry1.grid(column=1, row=2, pady=4)
    date_entry2.grid(column=2, row=2, pady=4)

    label3.grid(column=1, row=3, pady=4)
    int1.grid(column=2, row=3, pady=4)
    label4.grid(column=1, row=4, pady=4)
    int2.grid(column=2, row=4, pady=4)
    label5.grid(column=1, row=5, pady=4)
    int3.grid(column=2, row=5, pady=4)

    label6.grid(column=1, row=1, pady=4)
    int4.grid(column=1, row=2, pady=4)
    check_button_ASUTP.grid(column=1,row=3,pady=4)
    label7.grid(column=2, row=1, pady=4)
    int5.grid(column=2, row=2, pady=4)
    check_button_ASUE.grid(column=2,row=3,pady=4)
    label8.grid(column=3, row=1, pady=4)
    int6.grid(column=3, row=2, pady=4)
    check_button_TVS.grid(column=3,row=3,pady=4)
    label9.grid(column=4, row=1, pady=4)
    int7.grid(column=4, row=2, pady=4)
    check_button_ASPS.grid(column=4,row=3,pady=4)

    button1.grid(column=1, row=4, pady=4)
    button2.grid(column=2, row=4, pady=4)
    button3.grid(column=3, row=4, pady=4)
    button4.grid(column=4, row=4, pady=4)

    button9.grid(column=1, row=5, pady=4)
    button10.grid(column=2, row=5, pady=4)
    button11.grid(column=3, row=5, pady=4)
    button12.grid(column=4, row=5, pady=4)

    button5.grid(column=1, row=6, pady=4)
    button6.grid(column=2, row=6, pady=4)
    button7.grid(column=3, row=6, pady=4)
    button8.grid(column=4, row=6, pady=4)

    dir_button_ASUTP.grid(column=1, row=7, padx=4, pady=4)
    dir_button_ASUE.grid(column=2, row=7, padx=4, pady=4)
    dir_button_TVS.grid(column=3, row=7, padx=4, pady=4)
    dir_button_ASPS.grid(column=4, row=7, padx=4, pady=4)

    fileopen_button.pack()
    start_button.pack()
    check_button.pack(pady=4)

    frame1.pack()
    frame2.pack()
    frame3.pack()

    if config.getboolean('Start', 'autostart'):
        start(date_format, float(int1.get()), float(int2.get()), float(int3.get()), date_entry1.get_date(), date_entry2.get_date())




# Запуск главного цикла
    root.mainloop()


if __name__ == "__main__":
    main()
